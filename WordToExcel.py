import docx
import os
from openpyxl import Workbook
from tkinter import *
from tkinter import filedialog
from tkinter import ttk


#打开word文件夹
def OpenDictionary():
    Folderpath = filedialog.askdirectory()  # 获得选择好的文件夹
    print(Folderpath)
    #Filepath = filedialog.askopenfilename()  # 获得选择好的文件
    v1.set(Folderpath)
    path = '"'+ str(v1) + '"'
    print(path)

def wordToexcel():
    # 新建表格
    header = v2.split()
    location = v3.split()
    workbook = Workbook()
    booksheet = workbook.active
    for i in range(0,len(header)-1):
        booksheet.append(header[i])
    path ='"'+v1+'"'
    for file in os.walk(path):

        count = 0

        for filename in file[2]:
            if filename.split('.')[1] == "docx":
                count += 1
                print(str(count) + '、' + filename)
                f = docx.Document(path + filename)
                t = f.tables[0]
                list = []
                list.append(count)
                list.append(t.cell(0, 3).text)
                list.append(t.cell(0, 12).text)
                list.append(t.cell(1, 3).text)
                list.append(t.cell(1, 7).text)
                list.append(t.cell(1, 10).text)
                list.append(t.cell(1, 15).text)
                list.append(t.cell(2, 4).text)
                list.append(t.cell(2, 10).text)
                list.append(t.cell(7, 3).text)
                list.append(t.cell(7, 5).text)
                list.append(t.cell(7, 11).text)
                list.append(t.cell(13, 3).text)
                list.append(t.cell(14, 3).text)
                list.append(t.cell(15, 3).text)
                list.append(t.cell(17, 3).text)
                list.append(t.cell(18, 3).text)
                list.append(t.cell(18, 11).text)
                list.append(t.cell(19, 3).text)
                list.append(t.cell(20, 3).text)
                list.append(t.cell(21, 3).text)
                booksheet.append(list)
        workbook.save('D:\\工作\\省厅核查\\export.xlsx')
if __name__ == '__main__':
    root1 = Tk()
    root1.title('导入文件与模板')
    root1.geometry('900x750')
    root1.resizable(0, 0)
    v1 = StringVar()

    l1 = Label(root1, textvariable=v1,
               relief='groove', justify='left', bg='white', anchor='w')
    l1.place(x=250, y=25, width=270, height=30)
    # e1 = Entry(root,textvariable=v1)
    #

    b1 = Button(root1, text='打开word文件夹', command=OpenDictionary)
    b1.place(x=540, y=20, width=130, height=40)
    b2 = Button(root1, text='开始转换', command=wordToexcel)
    b2.place(x=400, y=650, width=100, height=40)
    # root.withdraw()
    # fpath = filedialog.askopenfilename();print(fpath)

    l2 = Label(root1, text='按顺序输入生成的Excel文件的属性（按空格隔开）')
    l2.place(x=340, y=75, height=50)
    v2 = StringVar()
    e1 = Entry(root1, textvariable=v2)
    e1.place(x=200, y=150, width=500, height=100)

    l3 = Label(root1, text='输入word表格中要提取的元素位置')
    l3.place(x=340, y=270, height=50)
    v3 = StringVar()
    e2 = Entry(root1, textvariable=v3)
    e2.place(x=200, y=350, width=500, height=100)
    root1.mainloop()